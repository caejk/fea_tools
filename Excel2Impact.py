#!/usr/bin/env python3

import openpyxl
import pandas as pd
import sys
import math
import os
'''
This script will help you in comparing your FEA simulation with real world crash test.
The impact points from a real world crash test are usually documented in an Excel.
This script produces Animator session file that will create impact points from the coordinates
saved in the Excel.
The idea is that you load your simulation in one Animator slot and read the session file 
created by this script. The session file will generate impact points in the pre-crash position
on the fea model in the first slot. It will also generate a second slot that will contain
the points from the real life crash test in their post-crash position.

Your data in the Excel sheet have to be sorted this way:
          A    |   B   |   C    |   D   |   E   |    F   |   G   |   H   |   I    |   J   | 
  1            |                        |                        |                        |
  2 ___________|________________________|________________________|________________________|
  3    Point   |           X            |            Y           |           Z            |
  4      Nr.   | before | after | delta | before | after | delta | before | after | delta |
  5       1       100       93     -7      120      116     -4       140     152     12 
  6       2      2400     2432     32       -5       2       7      1620    1595    -25   
  .
  ..
  ...

Console usage:
python3 Excel2Impact.py File.xlsm

TO DO: Feature to allow user defined rows and columns selection is in development..
'''
def UserInput():                             # User enters an Excel file as an argument in the console
    ExcelFile = sys.argv[1]
    wb = openpyxl.load_workbook(ExcelFile)
    sheet_list = wb.sheetnames
    print('-----------------')
    print('Available Sheets:')
    print('-----------------')
    print(*sheet_list, sep=', ')
    print('-----------------')
    wb.close()
    SheetName = input("Enter Sheet Name: ")  # User is given a list of Excel sheets to choose from
    return ExcelFile, SheetName
def ReadExcelData(ExcelFile, SheetName):     # Function for serializing data points from Excel
    wb = openpyxl.load_workbook(ExcelFile)
    df = pd.DataFrame(wb[SheetName].values)
    #    print(df)
    #    print(df.iloc[4:, 0])
    counter = 4
    for cell in df.iloc[4:, 0]:              # Determine maximum row with non-empty data
        #        if cell != None:
        content = str(cell)
        if content.isdigit():
            counter += 1

    MaxRow = counter                         # Sort pre- and post-crash point coordinates into dataframes
    df_points_coord_orig = df.loc[4:MaxRow, [0, 1, 4, 7]]
    df_points_coord_crash = df.loc[4:MaxRow, [0, 2, 5, 8]]
    df_points_coord_orig.rename({0: 'PointNr', 1: 'X', 4: 'Y', 7: 'Z'}, axis=1, inplace=True)
    df_points_coord_crash.rename({0: 'PointNr', 2: 'X', 5: 'Y', 8: 'Z'}, axis=1, inplace=True)
    MaxRowSortedOri = len(df_points_coord_orig)
    MaxRowSortedCra = len(df_points_coord_crash)
    df_points_coord_orig.reset_index(drop=True, inplace=True)
    df_points_coord_crash.reset_index(drop=True, inplace=True)
    df_points_coord_orig.index = pd.RangeIndex(start=1, stop=MaxRowSortedOri + 1, step=1)
    df_points_coord_crash.index = pd.RangeIndex(start=1, stop=MaxRowSortedCra + 1, step=1)

    #    print("Messpunkte bevor crash")
    #    print(df_points_coord_orig)
    #    print("Messpunkte nach crash")
    #    print(df_points_coord_crash)
    return df_points_coord_orig, df_points_coord_crash
def CreatePointObjects(df_ori, df_cra):    # Generates list of Point object
    ListOfPointObjects = []
    for i in range(1, len(df_ori)):
        Nr, X, Y, Z, Xcr, Ycr, Zcr = None, None, None, None, None, None, None
        Nr = df_ori.loc[i, 'PointNr']
        X = df_ori.loc[i, 'X']
        Y = df_ori.loc[i, 'Y']
        Z = df_ori.loc[i, 'Z']
        Xcr = df_cra.loc[i, 'X']
        Ycr = df_cra.loc[i, 'Y']
        Zcr = df_cra.loc[i, 'Z']
        Pt = Point(Nr, X, Y, Z, Xcr, Ycr, Zcr)
        ListOfPointObjects.append(Pt)

    return ListOfPointObjects
class Point:
    def __init__(self, Nr, X, Y, Z, Xcr, Ycr, Zcr):     # Initiate point with ID, pre- and post-crash coordinates
        self.Nr = Nr
        self.X = X
        self.Y = Y
        self.Z = Z

        if Xcr is not None:
            self.Xcr = Xcr
        else:
            self.Xcr = X
        if Ycr is not None:
            self.Ycr = Ycr
        else:
            self.Ycr = Y
        if Zcr is not None:
            self.Zcr = Zcr
        else:
            self.Zcr = Z

        if Xcr == None or Ycr == None or Zcr == None:
            self.message = 'Data Error'
        else:
            self.message = 'Data OK'
    def CheckForMessage(self):             # Print warning in case of missing post-crash coordinate
        if self.message == 'Data OK':
            pass
        elif self.message == 'Data Error':
            print('INFO: Point number ' + str(self.Nr) + ' has no data after crash test')
        else:
            print('INFO: Point number ' + str(self.Nr) + ' data status unclear')
    def GetCoord(self):
        Coord = [self.X, self.Y, self.Z]
        return Coord
    def GetDynaCoord(self):              # Produce Nodes in LS-Dyna syntax for pre-crash coordinates
        space8 = '        '
        space16 = '                '
        nodend = '       0       0'
        NodeID = str(self.Nr)
        NodeX = str(self.X)
        NodeY = str(self.Y)
        NodeZ = str(self.Z)
        lenspaceID = 8 - len(NodeID)
        lenspaceX = 16 - len(NodeX)
        lenspaceY = 16 - len(NodeY)
        lenspaceZ = 16 - len(NodeZ)
        DynaCoord = NodeID + ' ' * lenspaceID + ' ' * lenspaceX + NodeX + ' ' * lenspaceY + NodeY + ' ' * lenspaceZ + NodeZ + nodend
        return DynaCoord
    def GetDynaCoordCrashed(self):      # Produce Nodes in LS-Dyna syntax for post-crash coordinates
        space8 = '        '
        space16 = '                '
        nodend = '       0       0'
        NodeID = str(self.Nr)
        NodeX = str(self.Xcr)
        NodeY = str(self.Ycr)
        NodeZ = str(self.Zcr)
        lenspaceID = 8 - len(NodeID)
        lenspaceX = 16 - len(NodeX)
        lenspaceY = 16 - len(NodeY)
        lenspaceZ = 16 - len(NodeZ)
        DynaCoordCra = NodeID + ' ' * lenspaceID + ' ' * lenspaceX + NodeX + ' ' * lenspaceY + NodeY + ' ' * lenspaceZ + NodeZ + nodend
        return DynaCoordCra
    def GetDelta(self):                 # Calculate total displacement from pre- to post-crash
        try:
            dx = self.Xcr - self.X
            dy = self.Ycr - self.Y
            dz = self.Zcr - self.Z
            dtot = math.sqrt(pow(dx, 2) + pow(dy, 2) + pow(dz, 2))
            dtot = round(dtot, 1)
        except:
            dx, dy, dz, dtot = 0, 0, 0, 0
            print('$Point number ' + str(self.Nr) + ' delta not measurable after crash')
        return dx, dy, dz, dtot
    def CreateImpactPoint(self):        # Produce impact point in Animator syntax for pre-crash
        dx, dy, dz, dtot = self.GetDelta()
        CreateImp = 'imp usr cre xyz' + ' ' + str(self.X) + ' ' + str(self.Y) + ' ' + str(self.Z) + ' Number ' + str(
            self.Nr)
        #        print(CreateImp)
        #       slot.executeCommand(CreateImp)
        return CreateImp
    def CreateImpactPointCrashed(self):      # Produce impact point in Animator syntax for post-crash
        dx, dy, dz, dtot = self.GetDelta()
        CreateImpCra = 'imp usr cre xyz' + ' ' + str(self.Xcr) + ' ' + str(self.Ycr) + ' ' + str(
            self.Zcr) + ' Number ' + str(self.Nr) + ' dt: ' + str(dtot) + ' mm (dx ' + str(dx) + ' dy ' + str(
            dy) + ' dz ' + str(dz) + ')'
        #        print(CreateImpCra)
        return CreateImpCra

if __name__ == "__main__":
    ExcelFile, SheetName = UserInput()
    df_ori, df_cra = ReadExcelData(ExcelFile, SheetName)
    ListOfPointObjects = CreatePointObjects(df_ori, df_cra)

    if ' ' in SheetName: SheetName = SheetName.replace(' ', '_')

    with open('Messpunkte_' + SheetName + '_Simulation.ses', 'w') as s:       # For Simulation:
        print('$--------------------------------', file=s)                    # Write Animator Session File with pre-crash impact points
        print('$ Simulation Slot (Before crash):', file=s)
        print('$--------------------------------', file=s)
        for Pt in ListOfPointObjects:
            Pt.CreateImpactPoint()
            print(Pt.CreateImpactPoint(), file=s)

    with open('Messpunkte_' + SheetName + '_Versuch_Crashed.ses', 'w') as s:  # For Crash Test:
        print('$----------------------------', file=s)                        # Write Animator Session File with post-crash points
        print('$ Versuch Slot (After crash):', file=s)
        print('$----------------------------', file=s)
        print('$rea mod ...', file=s)
        for Pt in ListOfPointObjects:
            Pt.CreateImpactPoint()
            print(Pt.CreateImpactPointCrashed(), file=s)

    with open('Messpunkte_' + SheetName + '_Versuch_Crashed.key', 'w') as s:   # For Crash Test:
        print('$-------------------------------------------', file=s)          # Write LS-Dyna nodes with post-crash coordinates as key file
        print('$ Versuch Slot: points after crash as nodes:', file=s)
        print('$-------------------------------------------', file=s)
        print('*KEYWORD', file=s)
        print('*NODE', file=s)
        for Pt in ListOfPointObjects:
            Pt.CreateImpactPoint()
            print(Pt.GetDynaCoordCrashed(), file=s)
        print('*END', file=s)

    for Pt in ListOfPointObjects:
        Pt.CheckForMessage()
    cwd = os.getcwd()
    NodeDynaFile = cwd + '/Messpunkte_' + SheetName + '_Versuch_Crashed.key'

    with open('Messpunkte_' + SheetName + '.ses', 'w') as s:                # Generate main Animator session file
        print('s[all]:rea ses ' + cwd + '/Messpunkte_' + SheetName + '_Simulation.ses', file=s)
        print(
            's[new]:rea fil "Dyna3d" ' + cwd + '/Messpunkte_' + SheetName + '_Versuch_Crashed.key GEO=0:pid:all ADD=no',
            file=s)
        print('s["' + NodeDynaFile + '"]:rea ses ' + cwd + '/Messpunkte_' + SheetName + '_Versuch_Crashed.ses', file=s)
        print('s["' + NodeDynaFile + '"]:imp usr co1 magenta all', file=s)

    print('-----------------')
    print('Step 1: Open your validation simulation in animator')
    print('Step 2: rea ses ' + cwd + '/Messpunkte_' + SheetName + '.ses')
    print('-----------------')